"""
CSV/Excel parser for email data
"""
import csv
import os
from typing import List, Dict, Optional
from openpyxl import load_workbook
from config import COLUMN_ALIASES


class CSVParser:
    """Parse CSV/Excel files for email sending"""
    
    def __init__(self):
        self.data = []
        self.column_mapping = {}
        self.error_message = None
    
    def parse_file(self, file_path: str) -> bool:
        """
        Parse CSV or Excel file
        
        Args:
            file_path: Path to the data file
            
        Returns:
            True if successful, False otherwise
        """
        self.data = []
        self.column_mapping = {}
        self.error_message = None
        
        if not os.path.exists(file_path):
            self.error_message = f"File not found: {file_path}"
            return False
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_ext in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            elif file_ext == '.csv':
                return self._parse_csv(file_path)
            else:
                self.error_message = f"Unsupported file type: {file_ext}"
                return False
        except Exception as e:
            self.error_message = f"Error parsing file: {str(e)}"
            return False
    
    def _detect_delimiter(self, file_path: str) -> str:
        """Detect CSV delimiter (comma, semicolon, or tab)"""
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            first_line = f.readline()
            
        # Count occurrences of each potential delimiter
        comma_count = first_line.count(',')
        semicolon_count = first_line.count(';')
        tab_count = first_line.count('\t')
        
        # Return the most common delimiter
        if semicolon_count > comma_count and semicolon_count > tab_count:
            return ';'
        elif tab_count > comma_count and tab_count > semicolon_count:
            return '\t'
        else:
            return ','
    
    def _parse_csv(self, file_path: str) -> bool:
        """Parse CSV file with auto-detected delimiter"""
        try:
            # Detect delimiter
            delimiter = self._detect_delimiter(file_path)
            
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                rows = list(reader)
                
            if not rows:
                self.error_message = "CSV file is empty"
                return False
            
            return self._process_rows(rows, file_path)
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                delimiter = self._detect_delimiter(file_path)
                with open(file_path, 'r', encoding='latin-1') as f:
                    reader = csv.DictReader(f, delimiter=delimiter)
                    rows = list(reader)
                return self._process_rows(rows, file_path)
            except Exception as e:
                self.error_message = f"Encoding error: {str(e)}"
                return False
    
    def _parse_excel(self, file_path: str) -> bool:
        """Parse Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Get headers
            headers = [cell.value for cell in sheet[1]]
            if headers and None in headers:
                # Find last non-None column
                last_col = 0
                for i, h in enumerate(headers):
                    if h is not None:
                        last_col = i
                    else:
                        break
                headers = headers[:last_col + 1]
            
            if not headers or all(h is None for h in headers):
                self.error_message = "Excel file has no headers"
                return False
            
            # Convert to list of dicts
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and any(cell is not None for cell in row):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and header is not None:
                            row_dict[header] = row[i] if row[i] is not None else ""
                    rows.append(row_dict)
            
            wb.close()
            
            if not rows:
                self.error_message = "Excel file has no data rows"
                return False
            
            return self._process_rows(rows, file_path)
        except Exception as e:
            self.error_message = f"Error reading Excel file: {str(e)}"
            return False
    
    def _process_rows(self, rows: List[Dict], file_path: str) -> bool:
        """Process rows and build column mapping"""
        if not rows:
            self.error_message = "No data rows found"
            return False
        
        # Build column mapping from first row
        first_row = rows[0]
        self._build_column_mapping(first_row)
        
        # Validate required columns
        if "recipient_email" not in self.column_mapping:
            self.error_message = (
                "Missing required column. Please include one of: "
                + ", ".join(COLUMN_ALIASES["recipient_email"])
            )
            return False
        
        # Parse all rows
        for i, row in enumerate(rows, start=2):  # Start at row 2 (1-indexed)
            parsed_row = self._parse_row(row)
            if parsed_row:
                parsed_row['_row_num'] = i
                self.data.append(parsed_row)
        
        if not self.data:
            self.error_message = "No valid data rows found"
            return False
        
        return True
    
    def _build_column_mapping(self, row: Dict):
        """Map actual column names to standard names"""
        for standard_name, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in row:
                    self.column_mapping[standard_name] = alias
                    break
    
    def _parse_attachment_string(self, attachment_str: str) -> List[str]:
        """
        Parse attachment string that may contain multiple files.
        Supports comma-separated paths and quoted paths with commas.
        
        Examples:
            "file1.pdf" -> ["file1.pdf"]
            "file1.pdf,file2.pdf" -> ["file1.pdf", "file2.pdf"]
            '"C:\\My Documents,file.pdf",file2.pdf' -> ["C:\\My Documents,file.pdf", "file2.pdf"]
        """
        if not attachment_str or not attachment_str.strip():
            return []
        
        import shlex
        try:
            # Use shlex to properly handle quoted strings
            parsed = shlex.split(attachment_str.strip())
            return [p.strip() for p in parsed if p.strip()]
        except:
            # Fallback: simple comma split
            return [p.strip() for p in attachment_str.split(',') if p.strip()]
    
    def _parse_row(self, row: Dict) -> Optional[Dict]:
        """Parse a single row into standard format"""
        parsed = {}
        
        # Recipient email (required)
        recipient_col = self.column_mapping.get("recipient_email")
        if not recipient_col or recipient_col not in row:
            return None
        
        recipient = row[recipient_col]
        if not recipient or not str(recipient).strip():
            return None
        
        parsed["recipient_email"] = str(recipient).strip()
        
        # Subject (optional, can be from CSV or default)
        if "subject" in self.column_mapping:
            subject_col = self.column_mapping["subject"]
            parsed["subject"] = str(row[subject_col]).strip() if row.get(subject_col) else ""
        else:
            parsed["subject"] = ""
        
        # Attachment (optional) - now supports multiple files
        if "attachment_filename" in self.column_mapping:
            attachment_col = self.column_mapping["attachment_filename"]
            attachment_str = str(row[attachment_col]).strip() if row.get(attachment_col) else ""
            parsed["attachment_filenames"] = self._parse_attachment_string(attachment_str)
        else:
            parsed["attachment_filenames"] = []
        
        # Body content (optional)
        if "body_content" in self.column_mapping:
            body_col = self.column_mapping["body_content"]
            parsed["body_content"] = str(row[body_col]).strip() if row.get(body_col) else ""
        else:
            parsed["body_content"] = ""
        
        return parsed
    
    def get_data(self) -> List[Dict]:
        """Get parsed data"""
        return self.data
    
    def get_row_count(self) -> int:
        """Get number of parsed rows"""
        return len(self.data)
    
    def get_error(self) -> Optional[str]:
        """Get error message if parsing failed"""
        return self.error_message


def generate_template_csv(output_path: str) -> bool:
    """
    Generate a template CSV file with correct headers
    
    Args:
        output_path: Path where template will be saved
        
    Returns:
        True if successful
    """
    try:
        headers = [
            "recipient_email",
            "subject",
            "attachment_filename",
            "body_content"
        ]
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            # Add example rows
            writer.writerow([
                "recipient@example.com",
                "Example Subject",
                "/path/to/attachment.pdf",
                "This is the email body content."
            ])
            # Add example with multiple attachments
            writer.writerow([
                "another@example.com",
                "Multiple Attachments",
                '"file1.pdf,file2.pdf,file3.pdf"',
                "Email with multiple attachments."
            ])
        
        return True
    except Exception as e:
        print(f"Error generating template: {e}")
        return False
